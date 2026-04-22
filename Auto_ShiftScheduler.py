# %%
import pandas as pd
import random
import re
import os
import sys
import traceback
from collections import defaultdict
from openpyxl import load_workbook
from datetime import datetime

# ===================== 程序初始化（自动获取运行目录） =====================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)      # EXE 所在目录
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # 脚本所在目录

# 日志文件路径
LOG_FILE = os.path.join(BASE_DIR, "运行日志.txt")

def log_print(message, also_print=True):
    """同时将信息输出到控制台和日志文件"""
    if also_print:
        print(message)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(message + "\n")

# 写入日志头部（每次运行覆盖旧日志）
with open(LOG_FILE, "w", encoding="utf-8") as f:
    f.write(f"排班系统运行日志 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("=" * 50 + "\n")
def load_seed_from_config(base_dir):
    """从 config.txt 读取种子，若没有或格式错误则返回默认值"""
    config_path = os.path.join(base_dir, "config.txt")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                seed_str = f.read().strip()
                return int(seed_str)
        except:
            pass
    return 18888888  # 默认种子
# ===================== 核心配置 =====================
# 随机种子（如需每次运行结果不同，可改为 random.seed(int(time.time()))）
SEED = load_seed_from_config(BASE_DIR)
random.seed(SEED)
log_print(f"当前随机种子：{SEED}")
pd.set_option('future.no_silent_downcasting', True)

# 文件路径（全部基于 BASE_DIR）
# FREE_CLASS_FILE = os.path.join(BASE_DIR, "无课表.xlsx")          # 用户需将无课表重命名为此
# ===================== 自动查找无课表文件 =====================
def find_free_class_file(base_dir):
    """在 base_dir 目录下查找包含“无课表”字样的 .xlsx 文件"""
    for file in os.listdir(base_dir):
        if file.endswith('.xlsx') and '无课表' in file:
            return os.path.join(base_dir, file)
    return None

FREE_CLASS_FILE = find_free_class_file(BASE_DIR)
if FREE_CLASS_FILE is None:
    log_print("❌ 错误：在当前文件夹内找不到包含“无课表”字样的 .xlsx 文件。")
    input("按回车键退出...")
    sys.exit(1)
else:
    log_print(f" 已找到无课表文件：{os.path.basename(FREE_CLASS_FILE)}")

BASE_TEMPLATE_FILE = os.path.join(BASE_DIR, "moban.xlsx")        # 万年基础模板
OUTPUT_FILE = os.path.join(BASE_DIR, "值班表_输出.xlsx")         # 最终排班结果


# 固定规则
MAX_SHIFT_PER_PERSON = 1          # 每人最多值班次数
EXCLUDE_DEPART = "统筹部"         # 统筹部名称

# ===================== 工具函数：从文件名提取周次 =====================
def chinese_to_arabic(chinese_str):
    chinese_digits = {
        "零": 0, "一": 1, "二": 2, "三": 3, "四": 4,
        "五": 5, "六": 6, "七": 7, "八": 8, "九": 9
    }
    units = {"十": 10, "百": 100}
    total = 0
    temp = 0
    for char in chinese_str:
        if char in chinese_digits:
            temp = chinese_digits[char]
        elif char in units:
            unit_value = units[char]
            if temp == 0:
                temp = 1
            total += temp * unit_value
            temp = 0
        else:
            return None
    total += temp
    return total

def extract_week_number_from_filename(file_path):
    filename = os.path.basename(file_path)
    match = re.search(r"第(\d+)周", filename)
    if match:
        return int(match.group(1))
    match_ch = re.search(r"第([一二三四五六七八九十百]+)周", filename)
    if match_ch:
        return chinese_to_arabic(match_ch.group(1))
    return None

# ===================== 工具函数：生成当周模板 =====================
def generate_weekly_template(base_template_path, output_template_path, week_num, unique_date_objects):
    try:
        wb = load_workbook(base_template_path)
        ws = wb.active
        if week_num is not None:
            ws['A1'].value = f"第{week_num}周小办公室值班表"
        start_col = 2
        date_row = 2
        for idx, dt in enumerate(unique_date_objects):
            col = start_col + idx
            if col > 6:
                break
            cell = ws.cell(row=date_row, column=col, value=dt)
            cell.number_format = 'm"月"d"日"'
        wb.save(output_template_path)
        log_print(f"已生成中间模板文件：{os.path.basename(output_template_path)}")
    except Exception as e:
        log_print(f"生成模板失败：{e}")
        raise

# ===================== 第一步：读取无课表 =====================
log_print("读取无课表ing...")
try:
    df_raw = pd.read_excel(FREE_CLASS_FILE, header=None, dtype=str)
except FileNotFoundError:
    log_print("错误：找不到文件“无课表.xlsx”,请确保它与本程序放在同一文件夹内。")
    input("按回车键退出...")
    sys.exit(1)
except Exception as e:
    log_print(f"读取无课表失败：{e}")
    input("按回车键退出...")
    sys.exit(1)

try:
    date_row = df_raw.iloc[0, 4:].copy().ffill()
    time_row = df_raw.iloc[1, 4:].copy().values
except IndexError:
    log_print("无课表格式错误：请检查是否标准格式")
    input("按回车键退出...")
    sys.exit(1)

formatted_dates = []
date_objects = []

for d in date_row:
    if pd.isna(d):
        formatted_dates.append("")
        date_objects.append(None)
        continue
    d_clean = str(d).replace('"', '').replace("：", ":")
    try:
        dt = pd.to_datetime(d_clean)
        formatted_dates.append(f"{dt.month}月{dt.day}日")
        date_objects.append(dt)
    except:
        formatted_dates.append(d_clean)
        date_objects.append(None)

time_columns = [f"{d}_{t}" for d, t in zip(formatted_dates, time_row)]
all_columns = ["部门", "姓名", "职务", "班级"] + time_columns

df = df_raw.iloc[2:].copy()
df.columns = all_columns
df["部门"] = df["部门"].ffill()
df = df.dropna(subset=["姓名"]).reset_index(drop=True)

unique_dates = list(dict.fromkeys([x.split("_")[0] for x in time_columns if "_" in x]))
unique_times = list(dict.fromkeys([x.split("_")[1] for x in time_columns if "_" in x]))
EARLIEST_TIME = unique_times[0] if unique_times else ""
NIGHT_TIME = unique_times[-1] if unique_times else ""

unique_date_objects = []
seen_dates = set()
for d_str, d_obj in zip(formatted_dates, date_objects):
    if d_str and d_str not in seen_dates and d_obj is not None:
        seen_dates.add(d_str)
        unique_date_objects.append(d_obj)

if len(unique_dates) < 5:
    log_print(f"注意,无课表只有 {len(unique_dates)} 天，可能影响强制空缺规则。")

FORCE_EMPTY_DATE = unique_dates[4] if len(unique_dates) >= 5 else ""
FORCE_EMPTY_TIME = unique_times[-1] if len(unique_times) >= 1 else ""
FORCED_EMPTY_SLOT = f"{FORCE_EMPTY_DATE}_{FORCE_EMPTY_TIME}"

# ===================== 第二步：生成当周模板 =====================
week_num = extract_week_number_from_filename(FREE_CLASS_FILE)
TEMPLATE_FILE = os.path.join(BASE_DIR, f"当周模板_第{week_num if week_num else 'X'}周.xlsx")
try:
    generate_weekly_template(BASE_TEMPLATE_FILE, TEMPLATE_FILE, week_num, unique_date_objects)
except FileNotFoundError:
    log_print("错误:找不到基础模板文件“moban.xlsx”,请确保它与本程序放在同一文件夹内。")
    input("按回车键退出...")
    sys.exit(1)

# ===================== 第三步：初始化统计 =====================
shift_count = defaultdict(int)
daily_tongchou = defaultdict(int)
schedule_dict = {}
personal_free_time = defaultdict(int)

for idx, row in df.iterrows():
    free_cnt = sum(1 for col in time_columns if row[col] == "无课")
    personal_free_time[row["姓名"]] = free_cnt

# ===================== 第四步：核心排班算法 =====================
def select_person(staff_df):
    if staff_df.empty:
        return ""
    staff_df = staff_df.sample(frac=1, random_state=random.randint(0, 10000)).reset_index(drop=True)
    staff_df = staff_df.sort_values(by="姓名", key=lambda x: x.map(personal_free_time), ascending=True, kind='mergesort')
    officers = staff_df[staff_df["职务"].str.contains("干事", na=False)]
    if not officers.empty:
        return officers.iloc[0]["姓名"]
    ministers = staff_df[staff_df["职务"].str.contains("部长", na=False)]
    if not ministers.empty:
        return ministers.iloc[0]["姓名"]
    return staff_df.iloc[0]["姓名"]

# 阶段一
log_print("正在安排统筹部人员（晚自习,早班）...")
for date in unique_dates:
    night_slot = f"{date}_{NIGHT_TIME}"
    if night_slot == FORCED_EMPTY_SLOT or night_slot not in time_columns:
        continue
    free_staff = df[(df[night_slot] == "无课") & (df["部门"] == EXCLUDE_DEPART)].copy()
    free_staff = free_staff[free_staff["姓名"].map(lambda x: shift_count.get(x, 0) < MAX_SHIFT_PER_PERSON)]
    selected = select_person(free_staff)
    if selected:
        schedule_dict[night_slot] = selected
        shift_count[selected] += 1
        daily_tongchou[date] += 1
    else:
        schedule_dict[night_slot] = ""

days_shuffled = unique_dates.copy()
random.shuffle(days_shuffled)
early_tongchou_count = 0

for date in days_shuffled:
    if early_tongchou_count >= 2:
        break
    early_slot = f"{date}_{EARLIEST_TIME}"
    if early_slot == FORCED_EMPTY_SLOT or early_slot not in time_columns:
        continue
    if early_slot in schedule_dict:
        continue
    free_staff = df[(df[early_slot] == "无课") & (df["部门"] == EXCLUDE_DEPART)].copy()
    free_staff = free_staff[free_staff["姓名"].map(lambda x: shift_count.get(x, 0) < MAX_SHIFT_PER_PERSON)]
    selected = select_person(free_staff)
    if selected:
        schedule_dict[early_slot] = selected
        shift_count[selected] += 1
        daily_tongchou[date] += 1
        early_tongchou_count += 1

for date in unique_dates:
    if daily_tongchou[date] >= 1:
        continue
    day_slots = [s for s in time_columns if s.startswith(date)]
    other_slots = [s for s in day_slots if NIGHT_TIME not in s and s != FORCED_EMPTY_SLOT]
    random.shuffle(other_slots)
    assigned = False
    for slot in other_slots:
        if slot in schedule_dict:
            continue
        free_staff = df[(df[slot] == "无课") & (df["部门"] == EXCLUDE_DEPART)].copy()
        free_staff = free_staff[free_staff["姓名"].map(lambda x: shift_count.get(x, 0) < MAX_SHIFT_PER_PERSON)]
        selected = select_person(free_staff)
        if selected:
            schedule_dict[slot] = selected
            shift_count[selected] += 1
            daily_tongchou[date] += 1
            assigned = True
            break
    if not assigned:
        log_print(f"{date} 无统筹部人员可用，正在处理。")

# 阶段二
log_print("准备剩余班次...")
remaining_slots = [s for s in time_columns if s not in schedule_dict]
if FORCED_EMPTY_SLOT in remaining_slots:
    remaining_slots.remove(FORCED_EMPTY_SLOT)
    schedule_dict[FORCED_EMPTY_SLOT] = ""
remaining_slots = [s for s in remaining_slots if NIGHT_TIME not in s]
random.shuffle(remaining_slots)

# 阶段三
log_print("安排其他部门人员ing...")
for slot in remaining_slots:
    free_staff = df[(df[slot] == "无课") & (df["部门"] != EXCLUDE_DEPART)].copy()
    free_staff = free_staff[free_staff["姓名"].map(lambda x: shift_count.get(x, 0) < MAX_SHIFT_PER_PERSON)]
    selected = select_person(free_staff)
    if selected:
        schedule_dict[slot] = selected
        shift_count[selected] += 1
    else:
        schedule_dict[slot] = ""

# 阶段四
log_print("检查并填补空缺ing...")
for slot in time_columns:
    if schedule_dict.get(slot, "") == "" and slot != FORCED_EMPTY_SLOT:
        free_all = df[(df[slot] == "无课") & (~df["姓名"].isin(schedule_dict.values()))]
        if NIGHT_TIME in slot:
            non_tongchou = free_all[free_all["部门"] != EXCLUDE_DEPART]
            if not non_tongchou.empty:
                selected = select_person(non_tongchou)
                if selected:
                    schedule_dict[slot] = selected
                    shift_count[selected] = shift_count.get(selected, 0) + 1
                    log_print(f" 晚自习破例：{slot} 安排非统筹部 {selected}")
                    continue
            tongchou_free = free_all[free_all["部门"] == EXCLUDE_DEPART]
            if not tongchou_free.empty:
                selected = select_person(tongchou_free)
                if selected:
                    schedule_dict[slot] = selected
                    shift_count[selected] = shift_count.get(selected, 0) + 1
                    daily_tongchou[slot.split("_")[0]] += 1
                    log_print(f"晚自习再破例：{slot} 复用统筹部的 {selected}")
                    continue
        else:
            tongchou_free = free_all[free_all["部门"] == EXCLUDE_DEPART]
            if not tongchou_free.empty:
                selected = select_person(tongchou_free)
                if selected:
                    schedule_dict[slot] = selected
                    shift_count[selected] = shift_count.get(selected, 0) + 1
                    daily_tongchou[slot.split("_")[0]] += 1
                    log_print(f" 破例安排：{slot} 原空缺，现安排统筹部的 {selected}")
                    continue
            if not free_all.empty:
                selected = select_person(free_all)
                if selected:
                    schedule_dict[slot] = selected
                    shift_count[selected] = shift_count.get(selected, 0) + 1
                    log_print(f" 二次兜底：{slot} 安排 {selected}（原空缺）")
# ===================== 第五步：导出 =====================
def export_to_template():
    try:
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active
        for r in range(3, 8):
            for c in range(2, 7):
                ws.cell(row=r, column=c, value="")
        col_map = {0: 2, 1: 3, 2: 4, 3: 5, 4: 6}
        row_map = {0: 3, 1: 4, 2: 5, 3: 6, 4: 7}
        fill_num = 0
        for date_idx, date in enumerate(unique_dates):
            for time_idx, time_slot in enumerate(unique_times):
                name = schedule_dict.get(f"{date}_{time_slot}", "")
                if name and date_idx in col_map and time_idx in row_map:
                    ws.cell(row=row_map[time_idx], column=col_map[date_idx], value=name)
                    fill_num += 1
        wb.save(OUTPUT_FILE)
        log_print("=" * 60)
        log_print(f"导出成功,最终文件路径：{os.path.basename(OUTPUT_FILE)}")
        log_print(f" 本次共排班 {fill_num} 人")
        tongchou_total = sum(1 for name, count in shift_count.items()
                             if count > 0 and df[df["姓名"] == name]["部门"].iloc[0] == EXCLUDE_DEPART)
        early_t_count = sum(1 for (s, name) in schedule_dict.items()
                            if name and EARLIEST_TIME in s and df[df["姓名"] == name]["部门"].iloc[0] == EXCLUDE_DEPART)
        log_print(f"统筹部指标：总安排 {tongchou_total} 人 | 早班占用 {early_t_count} 次")
        log_print("=" * 60)
    except Exception as e:
        log_print(f"导出失败：{e}")
        raise

export_to_template()
# 删除中间模板文件
if os.path.exists(TEMPLATE_FILE):
    os.remove(TEMPLATE_FILE)
    log_print(f"已清理中间文件：{os.path.basename(TEMPLATE_FILE)}")
log_print("排班已结束")
log_print(f"详细日志已保存至：{os.path.basename(LOG_FILE)}")
input("按回车键退出...")


